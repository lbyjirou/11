"""报价单导出 API - 基于模板填充"""
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pathlib import Path
from io import BytesIO
from datetime import datetime
from urllib.parse import quote
from pydantic import BaseModel

router = APIRouter(prefix="/api/export", tags=["导出"])

# 模板文件路径
TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "报价单模板.xlsx"


class ExportRequest(BaseModel):
    al_price: float
    collector: dict
    fin: dict
    tube: dict
    components: list
    mfg_cost: float
    freight: float
    material_cost: float
    component_cost: float
    profit: float
    final_price: float


@router.post("/excel", summary="导出报价单Excel")
def export_excel(data: ExportRequest):
    # 1. 加载模板
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # 2. 填充铝锭价 (H3)
    ws['H3'] = data.al_price

    # 3. 填充主要材料
    # 集流管 (行6): D6规格(仅预设时填充), E6单价, F6数量, G6定额
    if data.collector.get('spec'):
        ws['D6'] = data.collector.get('spec')
    ws['E6'] = data.collector.get('unit_price', '')
    ws['F6'] = data.collector.get('qty', '')
    ws['G6'] = data.collector.get('subtotal', '')

    # 翅片 (行7): E7单价, F7数量, G7有公式自动计算
    ws['E7'] = data.fin.get('unit_price', '')
    ws['F7'] = data.fin.get('qty', '')

    # 扁管 (行8): E8单价, F8数量, G8有公式自动计算
    ws['E8'] = data.tube.get('unit_price', '')
    ws['F8'] = data.tube.get('qty', '')

    # 4. 填充其他部件 (行9-17, 共9个部件)
    # 顺序: 边板、堵帽、隔板、芯体支架、进口压板、出口压板、储液罐组件、密封件、包装箱
    for i, comp in enumerate(data.components[:9]):
        row = 9 + i
        ws[f'F{row}'] = comp.get('qty', '')
        ws[f'G{row}'] = comp.get('subtotal', '')

    # 5. 填充汇总数据
    # G18 材料成本 - 有公式 =SUM(G6:G17), 不填
    # G19 制造费用 - 需要填充
    ws['G19'] = data.mfg_cost
    # G20 利润 - 有公式 =(G18+G19)*0.1, 不填
    # G21 运费 - 需要填充
    ws['G21'] = data.freight
    # G22 合计 - 有公式 =SUM(G18:G21), 不填

    # 6. 返回文件流
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"报价单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
